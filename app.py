# -*- coding: utf-8 -*-
"""
Streamlit App: TFP Restock Engine (Excel 2021-safe)

Run locally:
  pip install -r requirements.txt
  streamlit run app.py

What you get:
- Downloadable Excel model with tabs:
  FULL, SETTINGS (EN+GR), ADMIN, SUPPLIER, PO READY, INSTRUCTIONS
- Excel 2021 safe formulas (no dynamic arrays / LET / FILTER / XLOOKUP)

vNext changes (requested in this chat):
- PO READY: Internal code = Color SKU = Variant SKU without last 3 digits (first 8 digits)
  SKU structure: 5 master + 3 color + 3 size
- PO READY: add NOTES after color
- PO READY: add TOTAL UNITS column before PHOTO
- PO READY: print setup (fit all columns to 1 page wide + repeat header row)
- ADMIN: add Color SKU column (so PO READY SUMIFS can match without array tricks)
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import xlsxwriter


# ----------------------------
# Helpers
# ----------------------------

_NON_DIGITS = re.compile(r"\D")


def sku11(value) -> Optional[str]:
    """Extract 11-digit numeric SKU."""
    if value is None:
        return None
    try:
        if isinstance(value, (float, np.floating)) and np.isnan(value):
            return None
    except Exception:
        pass
    try:
        if isinstance(value, (float, np.floating)):
            value = int(value)
    except Exception:
        pass
    s = _NON_DIGITS.sub("", str(value))
    return s if len(s) == 11 else None


def to_float(x, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, (float, np.floating)) and np.isnan(x):
            return default
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def clean_color_attr(x) -> str:
    if x is None:
        return ""
    try:
        if isinstance(x, (float, np.floating)) and np.isnan(x):
            return ""
    except Exception:
        pass
    s = str(x).strip()
    s = re.sub(r"^\s*Χρώμα\s*:\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^\s*Color\s*:\s*", "", s, flags=re.IGNORECASE)
    return s.strip()


def clean_product_name(raw: object) -> str:
    """
    Best-effort product-name cleanup (remove SKU blocks and obvious size suffixes).
    """
    if raw is None:
        return ""
    try:
        if isinstance(raw, (float, np.floating)) and np.isnan(raw):
            return ""
    except Exception:
        pass

    s = str(raw).strip()

    # remove [SKU]
    s = re.sub(r"\[\s*\d{6,}\s*\]", "", s).strip()

    # common separators before attrs
    s = re.sub(r"\s*\|\s*", " ", s).strip()

    # remove size fragments (EN/GR)
    s = re.sub(r"(,?\s*(Size|Μέγεθος)\s*:\s*[^,]+)$", "", s, flags=re.IGNORECASE).strip()

    # remove trailing dash size (e.g. "Name - 38")
    s = re.sub(r"\s*-\s*\d{2,3}\s*$", "", s).strip()

    # collapse spaces
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def parse_days_window_from_filename(filename: str) -> Optional[int]:
    """Detect day window if filename contains "YYYY-MM-DD - YYYY-MM-DD" (inclusive)."""
    m = re.search(r"(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})", filename or "")
    if not m:
        return None
    try:
        d1 = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        d2 = datetime.strptime(m.group(2), "%Y-%m-%d").date()
        return (d2 - d1).days + 1
    except Exception:
        return None


def extract_odoo_delivered_qty(pivot_bytes: bytes) -> pd.DataFrame:
    """
    From Odoo Pivot export (sale.report), extract delivered qty per SKU.
    Assumes:
    - Column A has text containing "[SKU]"
    - Column B has delivered qty
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(pivot_bytes), data_only=True)
    ws = wb.active

    rows = []
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        qty = ws.cell(r, 2).value
        if not name or not isinstance(name, str):
            continue
        m = re.search(r"\[(\d+)\]", name)
        if not m:
            continue
        s = m.group(1)
        if len(s) != 11:
            continue
        rows.append((s, to_float(qty, 0.0)))

    if not rows:
        return pd.DataFrame(columns=["Variant SKU", "Odoo Qty Delivered"])

    df = pd.DataFrame(rows, columns=["Variant SKU", "Odoo Qty Delivered"])
    return df.groupby("Variant SKU", as_index=False)["Odoo Qty Delivered"].sum()


def require_columns(df: pd.DataFrame, required: List[str], label: str) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{label} is missing required columns: {missing}")


def pick_first_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Return first existing column name from candidates."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def parse_size_sort_key(x: str):
    """Sort sizes: numeric first, then alpha."""
    s = str(x).strip()
    try:
        return (0, float(s))
    except Exception:
        return (1, s)


# ----------------------------
# SETTINGS (Excel sheet)
# ----------------------------

@dataclass(frozen=True)
class SettingRow:
    key: str
    value: object
    desc_en: str
    desc_gr: str


def build_settings(days_window: int, overrides: Dict[str, object]) -> List[SettingRow]:
    def v(key, default):
        return overrides.get(key, default)

    rows: List[SettingRow] = [
        SettingRow(
            "sales_source", v("sales_source", "SHOPIFY"),
            "Sales source for demand: SHOPIFY, ODOO, or MAX (use the higher).",
            "Πηγή πωλήσεων για τη ζήτηση: SHOPIFY, ODOO ή MAX (κρατάει το μεγαλύτερο)."
        ),
        SettingRow(
            "season_mode", v("season_mode", "FW"),
            "Season context: FW, SS, or NEUTRAL (no season scaling).",
            "Σεζόν αναφοράς: FW, SS ή NEUTRAL (χωρίς season scaling)."
        ),
        SettingRow(
            "year_tag", v("year_tag", "2025-2026"),
            "Season year label used for EXACT match detection. If you put '2025', it matches any season text containing 2025.",
            "Ετικέτα χρονιάς για EXACT match. Αν βάλεις '2025' θα πιάσει ό,τι περιέχει 2025 μέσα στο κείμενο της σεζόν."
        ),
        SettingRow(
            "target_weeks", int(v("target_weeks", 10)),
            "Target stock coverage in weeks.",
            "Στόχος κάλυψης σε εβδομάδες."
        ),
        SettingRow(
            "lead_time_weeks", int(v("lead_time_weeks", 3)),
            "Supplier lead time (weeks). Used for reorder point.",
            "Lead time προμηθευτή (εβδομάδες). Χρησιμοποιείται στο reorder point."
        ),
        SettingRow(
            "safety_buffer", float(v("safety_buffer", 0.10)),
            "Safety buffer on top of demand (0.10 = +10%).",
            "Safety buffer πάνω στη ζήτηση (0.10 = +10%)."
        ),
        SettingRow(
            "rounding_step", int(v("rounding_step", 6)),
            "Round up suggested units to this step (e.g., 6 pairs).",
            "Στρογγυλοποίηση προς τα πάνω (π.χ. 6 → 6/12/18)."
        ),
        SettingRow(
            "moq", int(v("moq", 6)),
            "Minimum order quantity per SKU (applies only when restocking).",
            "MOQ ανά SKU (ισχύει μόνο όταν προτείνουμε restock)."
        ),
        # Advanced (kept in Excel, optional in UI)
        SettingRow("mult_exact", float(v("mult_exact", 1.20)),
                   "Advanced: Demand multiplier for EXACT season+year matches.",
                   "Advanced: Multiplier για EXACT season+year."),
        SettingRow("mult_partial", float(v("mult_partial", 1.00)),
                   "Advanced: Multiplier for same season family but not the main year.",
                   "Advanced: Multiplier για ίδια σεζόν, άλλη χρονιά."),
        SettingRow("mult_offseason", float(v("mult_offseason", 0.65)),
                   "Advanced: Multiplier for off-season.",
                   "Advanced: Multiplier εκτός σεζόν."),
        SettingRow("mult_unknown", float(v("mult_unknown", 0.90)),
                   "Advanced: Multiplier for unknown season.",
                   "Advanced: Multiplier για άγνωστη σεζόν."),
        SettingRow("days_window", int(days_window),
                   "Days in the sales window (auto). Used to convert sales to weekly demand.",
                   "Ημέρες δεδομένων (auto). Για μετατροπή σε πωλήσεις/εβδομάδα."),
        SettingRow("allow_override_outlet", int(v("allow_override_outlet", 0)),
                   "Advanced: If 0, OUTLET SKUs are forced to 0 even with override.",
                   "Advanced: Αν 0, τα OUTLET είναι πάντα 0, ακόμα κι αν βάλεις override."),
    ]
    return rows


# ----------------------------
# Core build (FULL)
# ----------------------------

def build_full_df(
    odoo_bytes: bytes,
    shopify_bytes: bytes,
    shopify_filename: str,
    pivot_bytes: bytes,
) -> Tuple[pd.DataFrame, int, Dict[str, int]]:

    odoo_df = pd.read_excel(BytesIO(odoo_bytes), dtype=str)
    shop_df = pd.read_csv(BytesIO(shopify_bytes), dtype=str)

    require_columns(odoo_df, ["Internal Reference"], "Odoo Product Variant export")
    require_columns(shop_df, ["Product variant SKU"], "Shopify sales export")

    pivot_df = extract_odoo_delivered_qty(pivot_bytes)
    days_window = parse_days_window_from_filename(shopify_filename) or 30

    base = odoo_df.copy()
    base["Variant SKU"] = base["Internal Reference"].apply(sku11)
    base = base[base["Variant SKU"].notna()].copy()
    if base.empty:
        raise ValueError("No valid 11-digit SKUs found in Odoo Product Variant file (Internal Reference).")

    # best-effort: product type + name + photo
    col_type = pick_first_col(base, [
        "Product Type", "Product type", "Type",
        "Product Category", "Product Category/Complete Name", "Product Category/Name"
    ])
    col_name = pick_first_col(base, ["Name", "Display Name", "Product Name", "Product Template", "Product Template/Name"])
    col_photo = pick_first_col(base, [
        "Image Url", "Image URL", "Image", "Photo", "PHOTO",
        "Image 128", "Image 1920", "Website Image", "Website Image/Url"
    ])

    product_type = base[col_type].astype(str).fillna("") if col_type else ""
    product_name_raw = base[col_name].astype(str).fillna("") if col_name else ""
    product_name = product_name_raw.apply(clean_product_name) if col_name else ""

    full = pd.DataFrame({
        "Product Type": product_type,
        "Product Name": product_name,
        "Variant SKU": base["Variant SKU"].astype(str),
        "Group Key": base.get("Color SKU", "").astype(str).fillna(""),
        "Our Code": base.get("Master Code", "").astype(str).fillna(""),
        "Color": base.get("Color Attribute", "").apply(clean_color_attr),
        "Size": base.get("Size Value", "").astype(str).fillna(""),
        "Vendor": base.get("Vendors/Display Name", "").astype(str).fillna(""),
        "Vendor Code": base.get("Vendors/Vendor Product Code", "").astype(str).fillna(""),
        "Product Season": base.get("Product Season", "").astype(str).fillna(""),
        "On Hand": base.get("Quantity On Hand", 0).apply(to_float),
        "Reserved / Outgoing": base.get("Outgoing Quantity", 0).apply(to_float) if "Outgoing Quantity" in base.columns else 0.0,
        "Incoming": base.get("Incoming Quantity", 0).apply(to_float) if "Incoming Quantity" in base.columns else 0.0,
        "Forecasted (Odoo)": base.get("Forecasted Quantity", "").apply(to_float, default=np.nan),
        "Photo": base[col_photo].astype(str).fillna("") if col_photo else "",
    })

    # Derived codes from Variant SKU (11 digits): 5 master + 3 color + 3 size
    # Color SKU (internal code) = Variant SKU without last 3 digits (first 8 digits)
    full["Color SKU"] = full["Variant SKU"].apply(lambda x: (str(x)[:-3] if isinstance(x, str) and len(x) == 11 else ""))

    # Shopify aggregate
    s = shop_df.copy()
    s["Variant SKU"] = s["Product variant SKU"].apply(sku11)
    s = s[s["Variant SKU"].notna()].copy()

    for c in ["Net sales", "Net items sold", "Cost of goods sold", "Gross profit", "Quantity ordered", "Quantity returned"]:
        if c in s.columns:
            s[c] = s[c].apply(to_float)
        else:
            s[c] = 0.0

    agg = s.groupby("Variant SKU", as_index=False)[
        ["Net sales", "Net items sold", "Cost of goods sold", "Gross profit", "Quantity ordered", "Quantity returned"]
    ].sum()

    agg["Gross margin"] = np.where(agg["Net sales"] > 0, agg["Gross profit"] / agg["Net sales"], 0.0)
    agg["Return rate"] = np.where(agg["Quantity ordered"] > 0, agg["Quantity returned"] / agg["Quantity ordered"], 0.0)
    agg.rename(columns={"Net items sold": "Shop Net items sold", "Cost of goods sold": "COGS"}, inplace=True)

    full = full.merge(
        agg[["Variant SKU", "Shop Net items sold", "Net sales", "COGS", "Gross profit", "Gross margin", "Return rate"]],
        on="Variant SKU",
        how="left",
    )

    for c in ["Shop Net items sold", "Net sales", "COGS", "Gross profit", "Gross margin", "Return rate"]:
        full[c] = full[c].fillna(0.0)

    full = full.merge(pivot_df, on="Variant SKU", how="left")
    full["Odoo Qty Delivered"] = full["Odoo Qty Delivered"].fillna(0.0)

    stats = {
        "warehouse_skus": int(full["Variant SKU"].nunique()),
        "shopify_skus_with_sales": int(full.loc[full["Shop Net items sold"] > 0, "Variant SKU"].nunique()),
        "odoo_skus_with_delivered": int(full.loc[full["Odoo Qty Delivered"] > 0, "Variant SKU"].nunique()),
        "skus_with_zero_sales": int(full.loc[(full["Shop Net items sold"] <= 0) & (full["Odoo Qty Delivered"] <= 0), "Variant SKU"].nunique()),
    }

    return full, days_window, stats


# ----------------------------
# Excel writer
# ----------------------------

def build_excel_bytes(full: pd.DataFrame, days_window: int, settings_overrides: Dict[str, object]) -> bytes:
    settings = build_settings(days_window, settings_overrides)
    n = len(full)
    full_cols = list(full.columns)

    sizes = sorted([s for s in full["Size"].astype(str).fillna("").unique().tolist() if str(s).strip() != ""], key=parse_size_sort_key)

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {"in_memory": True, "nan_inf_to_errors": True})

    ws_full = wb.add_worksheet("FULL")
    ws_set = wb.add_worksheet("SETTINGS")
    ws_admin = wb.add_worksheet("ADMIN")
    ws_sup = wb.add_worksheet("SUPPLIER")
    ws_po = wb.add_worksheet("PO READY")
    ws_help = wb.add_worksheet("INSTRUCTIONS")

    fmt_header = wb.add_format({"bold": True, "bg_color": "#F2F2F2", "border": 1})
    fmt_note = wb.add_format({"text_wrap": True})

    fmt_engine_header = wb.add_format({"bold": True, "bg_color": "#FFF2CC", "border": 1})
    fmt_engine_cell = wb.add_format({"bg_color": "#FFF2CC", "border": 1})

    # SETTINGS
    ws_set.write(0, 0, "key", fmt_header)
    ws_set.write(0, 1, "value", fmt_header)
    ws_set.write(0, 2, "description (EN)", fmt_header)
    ws_set.write(0, 3, "επεξήγηση (GR) + παράδειγμα", fmt_header)

    key_row: Dict[str, int] = {}
    for i, row in enumerate(settings, start=1):
        ws_set.write(i, 0, row.key)
        if isinstance(row.value, (int, float)):
            ws_set.write_number(i, 1, float(row.value))
        else:
            ws_set.write_string(i, 1, str(row.value))
        ws_set.write(i, 2, row.desc_en)
        ws_set.write(i, 3, row.desc_gr)
        key_row[row.key] = i + 1

    ws_set.set_column(0, 0, 26)
    ws_set.set_column(1, 1, 18)
    ws_set.set_column(2, 2, 70)
    ws_set.set_column(3, 3, 85)
    ws_set.freeze_panes(1, 0)

    ws_set.add_table(
        0, 0, len(settings), 3,
        {
            "name": "tbl_settings",
            "style": "Table Style Light 9",
            "autofilter": True,
            "columns": [
                {"header": "key"},
                {"header": "value"},
                {"header": "description (EN)"},
                {"header": "επεξήγηση (GR) + παράδειγμα"},
            ],
        },
    )

    def S(key: str) -> str:
        return f"SETTINGS!$B${key_row[key]}"

    # FULL values
    for j, c in enumerate(full_cols):
        ws_full.write(0, j, c, fmt_header)

    for i in range(n):
        for j, c in enumerate(full_cols):
            val = full.iloc[i, j]
            if isinstance(val, (int, float, np.integer, np.floating)) and not pd.isna(val):
                ws_full.write_number(i + 1, j, float(val))
            else:
                if pd.isna(val):
                    ws_full.write_blank(i + 1, j, None)
                else:
                    ws_full.write_string(i + 1, j, str(val))

    ws_full.freeze_panes(1, 0)
    ws_full.set_column(0, 6, 18)
    ws_full.set_column(7, len(full_cols) - 1, 16)

    ws_full.add_table(
        0, 0, n, len(full_cols) - 1,
        {
            "name": "tbl_full",
            "style": "Table Style Light 9",
            "autofilter": True,
            "columns": [{"header": c} for c in full_cols],
        },
    )

    # ADMIN
    admin_headers = [
        "Product Type", "Product Name",
        "Variant SKU", "Our Code", "Color", "Size",
        "Vendor", "Vendor Code", "Product Season",
        "On Hand", "Reserved / Outgoing", "Incoming",
        "Forecasted (Odoo)", "Effective Forecasted",
        "Engine Suggested", "Override Qty", "Final Qty",
        "Shop Net items sold", "Odoo Qty Delivered", "Net items sold (used)",
        "Units/week adj", "Demand units (target)", "Gap", "Reorder point",
        "Season Class", "Season Mult",
        "Return rate", "Gross margin", "Gross profit", "Net sales", "COGS",
        "Color SKU",
        "Reason",
    ]

    for j, h in enumerate(admin_headers):
        if h in ("Engine Suggested", "Override Qty", "Final Qty"):
            ws_admin.write(0, j, h, fmt_engine_header)
        else:
            ws_admin.write(0, j, h, fmt_header)

    full_col_index = {c: i for i, c in enumerate(full_cols)}

    def xlcol(idx: int) -> str:
        return xlsxwriter.utility.xl_col_to_name(idx)

    def F(colname: str, excel_row: int) -> str:
        return f"FULL!${xlcol(full_col_index[colname])}${excel_row}"

    # Column mapping (ADMIN) — important for PO READY SUMIFS
    # A Product Type
    # B Product Name
    # C Variant SKU
    # D Our Code (master code)
    # E Color
    # F Size
    # G Vendor
    # H Vendor Code
    # ...
    # Q Final Qty
    # AF Color SKU (internal code, 8 digits)
    # AG Reason

    for i in range(n):
        r = i + 2  # excel row

        ws_admin.write_formula(i + 1, 0, f"={F('Product Type', r)}")
        ws_admin.write_formula(i + 1, 1, f"={F('Product Name', r)}")
        ws_admin.write_formula(i + 1, 2, f"={F('Variant SKU', r)}")
        ws_admin.write_formula(i + 1, 3, f"={F('Our Code', r)}")
        ws_admin.write_formula(i + 1, 4, f"={F('Color', r)}")
        ws_admin.write_formula(i + 1, 5, f"={F('Size', r)}")
        ws_admin.write_formula(i + 1, 6, f"={F('Vendor', r)}")
        ws_admin.write_formula(i + 1, 7, f"={F('Vendor Code', r)}")
        ws_admin.write_formula(i + 1, 8, f"={F('Product Season', r)}")

        ws_admin.write_formula(i + 1, 9, f"={F('On Hand', r)}")
        ws_admin.write_formula(i + 1, 10, f"={F('Reserved / Outgoing', r)}")
        ws_admin.write_formula(i + 1, 11, f"={F('Incoming', r)}")
        ws_admin.write_formula(i + 1, 12, f"={F('Forecasted (Odoo)', r)}")

        season_cell = f"$I{r}"
        fw = f'OR(ISNUMBER(SEARCH("Φθινόπωρο",{season_cell})),ISNUMBER(SEARCH("Χειμώνας",{season_cell})),ISNUMBER(SEARCH("FW",{season_cell})),ISNUMBER(SEARCH("Fall",{season_cell})),ISNUMBER(SEARCH("Winter",{season_cell})))'
        ss = f'OR(ISNUMBER(SEARCH("Άνοιξη",{season_cell})),ISNUMBER(SEARCH("Καλοκαίρι",{season_cell})),ISNUMBER(SEARCH("SS",{season_cell})),ISNUMBER(SEARCH("Spring",{season_cell})),ISNUMBER(SEARCH("Summer",{season_cell})))'
        hy = f'IF({S("year_tag")}="",FALSE,ISNUMBER(SEARCH({S("year_tag")},{season_cell})))'
        season_class = (
            f'IF({S("season_mode")}="NEUTRAL","PARTIAL",'
            f'IF({S("season_mode")}="FW",'
            f'IF(AND({fw},{hy}),"EXACT",IF({fw},"PARTIAL",IF({ss},"OFFSEASON","UNKNOWN"))),'
            f'IF({S("season_mode")}="SS",'
            f'IF(AND({ss},{hy}),"EXACT",IF({ss},"PARTIAL",IF({fw},"OFFSEASON","UNKNOWN"))),'
            f'"UNKNOWN")))'
        )
        ws_admin.write_formula(i + 1, 24, f"={season_class}")
        ws_admin.write_formula(
            i + 1,
            25,
            f'=IF($Y{r}="EXACT",{S("mult_exact")},IF($Y{r}="PARTIAL",{S("mult_partial")},IF($Y{r}="OFFSEASON",{S("mult_offseason")},{S("mult_unknown")})))'
        )

        ws_admin.write_formula(i + 1, 13, f'=IF($M{r}<>"",$M{r},($J{r}-$K{r}+$L{r}))')

        ws_admin.write_formula(i + 1, 17, f"={F('Shop Net items sold', r)}")
        ws_admin.write_formula(i + 1, 18, f"={F('Odoo Qty Delivered', r)}")

        ws_admin.write_formula(i + 1, 19, f'=IF({S("sales_source")}="SHOPIFY",$R{r},IF({S("sales_source")}="ODOO",$S{r},MAX($R{r},$S{r})))')

        ws_admin.write_formula(i + 1, 26, f"={F('Return rate', r)}")
        ws_admin.write_formula(i + 1, 27, f"={F('Gross margin', r)}")
        ws_admin.write_formula(i + 1, 28, f"={F('Gross profit', r)}")
        ws_admin.write_formula(i + 1, 29, f"={F('Net sales', r)}")
        ws_admin.write_formula(i + 1, 30, f"={F('COGS', r)}")

        ws_admin.write_formula(i + 1, 20, f'=IF({S("days_window")}>0,($T{r}/{S("days_window")})*7*(1-$AA{r})*$Z{r},0)')
        ws_admin.write_formula(i + 1, 21, f'=$U{r}*{S("target_weeks")}*(1+{S("safety_buffer")})')
        ws_admin.write_formula(i + 1, 22, f'=$V{r}-$N{r}')
        ws_admin.write_formula(i + 1, 23, f'=$U{r}*{S("lead_time_weeks")}*(1+{S("safety_buffer")})')

        eng = f'=IF(ISNUMBER(SEARCH("OUTLET",$I{r})),0,IF($W{r}<=0,0,MAX({S("moq")},CEILING($W{r},{S("rounding_step")}))))'
        ws_admin.write_formula(i + 1, 14, eng, fmt_engine_cell)

        ws_admin.write_blank(i + 1, 15, None, fmt_engine_cell)

        ws_admin.write_formula(
            i + 1,
            16,
            f'=IF(AND(ISNUMBER(SEARCH("OUTLET",$I{r})),{S("allow_override_outlet")}=0),0,IF($P{r}<>"",$P{r},$O{r}))',
            fmt_engine_cell
        )

        # Color SKU (AF): Variant SKU without last 3 digits => first 8 digits
        ws_admin.write_formula(i + 1, 31, f'=IF(LEN($C{r})=11,LEFT($C{r},8),"")')

        # Reason (AG)
        ws_admin.write_formula(i + 1, 32, f'=IF(ISNUMBER(SEARCH("OUTLET",$I{r})),"OUTLET",IF($W{r}>0,"RESTOCK","NO_ACTION"))')

    ws_admin.freeze_panes(1, 4)
    ws_admin.set_column(0, 0, 18)
    ws_admin.set_column(1, 1, 34)
    ws_admin.set_column(2, 8, 16)
    ws_admin.set_column(9, 13, 14)
    ws_admin.set_column(14, 16, 14)
    ws_admin.set_column(17, 32, 16)

    ws_admin.add_table(
        0, 0, n, len(admin_headers) - 1,
        {
            "name": "tbl_admin",
            "style": "Table Style Light 9",
            "autofilter": True,
            "columns": [{"header": h} for h in admin_headers],
        },
    )

    # SUPPLIER (includes product name)
    sup_headers = [
        "vendor", "vendor code", "product name", "color", "size",
        "Qty", "our code", "variant sku", "notes", "image link"
    ]
    for j, h in enumerate(sup_headers):
        ws_sup.write(0, j, h, fmt_header)

    for i in range(n):
        r = i + 2
        ws_sup.write_formula(i + 1, 0, f"=ADMIN!$G{r}")
        ws_sup.write_formula(i + 1, 1, f"=ADMIN!$H{r}")
        ws_sup.write_formula(i + 1, 2, f"=ADMIN!$B{r}")
        ws_sup.write_formula(i + 1, 3, f"=ADMIN!$E{r}")
        ws_sup.write_formula(i + 1, 4, f"=ADMIN!$F{r}")
        ws_sup.write_formula(i + 1, 5, f"=ADMIN!$Q{r}")
        ws_sup.write_formula(i + 1, 6, f"=ADMIN!$D{r}")
        ws_sup.write_formula(i + 1, 7, f"=ADMIN!$C{r}")
        ws_sup.write_blank(i + 1, 8, None)
        # Photo from FULL (robust)
        if "Photo" in full_col_index:
            ws_sup.write_formula(i + 1, 9, f"=FULL!${xlcol(full_col_index['Photo'])}${r}")
        else:
            ws_sup.write_blank(i + 1, 9, None)

    ws_sup.freeze_panes(1, 0)
    ws_sup.set_column(0, 1, 22)
    ws_sup.set_column(2, 2, 36)
    ws_sup.set_column(3, 4, 14)
    ws_sup.set_column(5, 5, 10)
    ws_sup.set_column(6, 7, 16)
    ws_sup.set_column(8, 9, 28)

    ws_sup.add_table(
        0, 0, n, len(sup_headers) - 1,
        {
            "name": "tbl_supplier",
            "style": "Table Style Light 9",
            "autofilter": True,
            "columns": [{"header": h} for h in sup_headers],
        },
    )

    # PO READY (matrix)
    # Rows unique by (Vendor Code, Color SKU, Color, Photo)
    po_base = full[["Vendor Code", "Color SKU", "Color", "Photo"]].copy()
    po_base["Vendor Code"] = po_base["Vendor Code"].astype(str).fillna("")
    po_base["Color SKU"] = po_base["Color SKU"].astype(str).fillna("")
    po_base["Color"] = po_base["Color"].astype(str).fillna("")
    po_base["Photo"] = po_base["Photo"].astype(str).fillna("")
    po_base = po_base.drop_duplicates().reset_index(drop=True)

    po_headers = ["ΚΩΔΙΚΟΣ ΠΡΟΜΗΘΕΥΤΗ", "ΕΣΩΤΕΡΙΚΟΣ ΚΩΔΙΚΟΣ", "ΧΡΩΜΑ", "NOTES"] + sizes + ["ΣΥΝΟΛΟ ΤΜΧ", "PHOTO"]
    for j, h in enumerate(po_headers):
        ws_po.write(0, j, h, fmt_header)

    for i in range(len(po_base)):
        rr = i + 2
        ws_po.write_string(i + 1, 0, str(po_base.loc[i, "Vendor Code"]))
        ws_po.write_string(i + 1, 1, str(po_base.loc[i, "Color SKU"]))
        ws_po.write_string(i + 1, 2, str(po_base.loc[i, "Color"]))
        ws_po.write_blank(i + 1, 3, None)  # NOTES

        for j, size in enumerate(sizes):
            col_idx = 4 + j  # E onwards (after NOTES)
            header_cell = xlsxwriter.utility.xl_rowcol_to_cell(0, col_idx, row_abs=True, col_abs=True)

            # SUMIFS over ADMIN Final Qty:
            # Vendor Code in ADMIN column H
            # Color SKU in ADMIN column AF
            # Color in ADMIN column E
            # Size in ADMIN column F
            fml = (
                f"=SUMIFS(ADMIN!$Q:$Q,"
                f"ADMIN!$H:$H,$A{rr},"
                f"ADMIN!$AF:$AF,$B{rr},"
                f"ADMIN!$E:$E,$C{rr},"
                f"ADMIN!$F:$F,{header_cell})"
            )
            ws_po.write_formula(i + 1, col_idx, fml)

        # TOTAL units
        if len(sizes) > 0:
            first_cell = xlsxwriter.utility.xl_rowcol_to_cell(i + 1, 4)
            last_cell = xlsxwriter.utility.xl_rowcol_to_cell(i + 1, 4 + len(sizes) - 1)
            ws_po.write_formula(i + 1, 4 + len(sizes), f"=SUM({first_cell}:{last_cell})")
        else:
            ws_po.write_number(i + 1, 4, 0)

        # PHOTO last
        ws_po.write_string(i + 1, 5 + len(sizes), str(po_base.loc[i, "Photo"]))

    ws_po.freeze_panes(1, 4)
    ws_po.set_column(0, 2, 18)
    ws_po.set_column(3, 3, 22)  # NOTES
    ws_po.set_column(4, 4 + max(0, len(sizes) - 1), 6)
    ws_po.set_column(4 + len(sizes), 4 + len(sizes), 12)  # TOTAL
    ws_po.set_column(len(po_headers) - 1, len(po_headers) - 1, 55)  # PHOTO

    ws_po.add_table(
        0, 0, len(po_base), len(po_headers) - 1,
        {
            "name": "tbl_po_ready",
            "style": "Table Style Light 9",
            "autofilter": True,
            "columns": [{"header": h} for h in po_headers],
        },
    )

    # PO READY print setup (Excel -> Print / Save as PDF)
    ws_po.set_landscape()
    ws_po.set_paper(9)         # A4
    ws_po.fit_to_pages(1, 0)   # fit all columns to 1 page wide
    ws_po.repeat_rows(0)       # repeat header row on each page
    ws_po.set_margins(0.2, 0.2, 0.3, 0.3)

    # INSTRUCTIONS
    instr = (
        "INSTRUCTIONS (simple):\n"
        "1) Upload the 3 inputs and download the Excel.\n"
        "2) SETTINGS: focus on essentials.\n"
        "3) ADMIN: type Override Qty if needed (Final Qty updates).\n"
        "4) SUPPLIER and PO READY update automatically based on ADMIN Final Qty.\n"
        "Note: adding/removing SKUs/rows requires regenerating the Excel from the app."
    )
    ws_help.write(0, 0, instr, fmt_note)
    ws_help.set_column(0, 0, 130)

    wb.close()
    output.seek(0)
    return output.read()


# ----------------------------
# Streamlit UI
# ----------------------------

def main():
    st.set_page_config(page_title="TFP Restock Engine", layout="wide")

    st.title("TFP Restock Engine")
    st.caption("Build an Excel 2021-safe restock model (FULL / SETTINGS / ADMIN / SUPPLIER / PO READY).")

    colA, colB = st.columns([1.1, 0.9], gap="large")

    with colA:
        st.subheader("1) Upload inputs")
        odoo_file = st.file_uploader("Odoo Product Variant (product.product) .xlsx (warehouse truth set)", type=["xlsx"], key="odoo")
        shop_file = st.file_uploader("Shopify Net sales export .csv", type=["csv"], key="shop")
        pivot_file = st.file_uploader("Odoo Pivot Sales Analysis (sale.report) .xlsx", type=["xlsx"], key="pivot")

        st.subheader("2) Output")
        default_name = f"Restock_Model_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
        out_name = st.text_input("Output filename", value=default_name)

        st.divider()
        st.subheader("Settings (essentials)")
        s1, s2 = st.columns(2)

        with s1:
            sales_source = st.selectbox("sales_source", ["SHOPIFY", "ODOO", "MAX"], index=0)
            season_mode = st.selectbox("season_mode", ["FW", "SS", "NEUTRAL"], index=0)
            year_tag = st.text_input("year_tag", value="2025-2026")

        with s2:
            target_weeks = st.number_input("target_weeks", min_value=1, max_value=30, value=10, step=1)
            lead_time_weeks = st.number_input("lead_time_weeks", min_value=0, max_value=20, value=3, step=1)
            safety_buffer = st.number_input("safety_buffer", min_value=0.0, max_value=1.0, value=0.10, step=0.01, format="%.2f")
            rounding_step = st.number_input("rounding_step", min_value=1, max_value=60, value=6, step=1)
            moq = st.number_input("moq", min_value=0, max_value=60, value=6, step=1)

        with st.expander("Advanced (optional)", expanded=False):
            a1, a2 = st.columns(2)
            with a1:
                mult_exact = st.number_input("mult_exact", min_value=0.0, max_value=5.0, value=1.20, step=0.05, format="%.2f")
                mult_partial = st.number_input("mult_partial", min_value=0.0, max_value=5.0, value=1.00, step=0.05, format="%.2f")
            with a2:
                mult_offseason = st.number_input("mult_offseason", min_value=0.0, max_value=5.0, value=0.65, step=0.05, format="%.2f")
                mult_unknown = st.number_input("mult_unknown", min_value=0.0, max_value=5.0, value=0.90, step=0.05, format="%.2f")
            allow_override_outlet = st.selectbox("allow_override_outlet", [0, 1], index=0)

        settings_overrides = dict(
            sales_source=sales_source,
            season_mode=season_mode,
            year_tag=year_tag.strip(),
            target_weeks=int(target_weeks),
            lead_time_weeks=int(lead_time_weeks),
            safety_buffer=float(safety_buffer),
            rounding_step=int(rounding_step),
            moq=int(moq),
            mult_exact=float(mult_exact),
            mult_partial=float(mult_partial),
            mult_offseason=float(mult_offseason),
            mult_unknown=float(mult_unknown),
            allow_override_outlet=int(allow_override_outlet),
        )

        build_btn = st.button("Build Excel model", type="primary", use_container_width=True)

    with colB:
        st.subheader("Preview / Diagnostics")
        st.write("Upload the 3 inputs and click **Build Excel model**. You will get a download button.")

        if "last_error" in st.session_state and st.session_state["last_error"]:
            st.error(st.session_state["last_error"])

        if "stats" in st.session_state and st.session_state["stats"]:
            stats = st.session_state["stats"]
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Warehouse SKUs", stats.get("warehouse_skus", 0))
            k2.metric("Shopify SKUs with sales", stats.get("shopify_skus_with_sales", 0))
            k3.metric("Odoo SKUs delivered", stats.get("odoo_skus_with_delivered", 0))
            k4.metric("SKUs with zero sales", stats.get("skus_with_zero_sales", 0))

        if "full_preview" in st.session_state and isinstance(st.session_state["full_preview"], pd.DataFrame):
            st.write("Sample rows (FULL)")
            st.dataframe(st.session_state["full_preview"], use_container_width=True, height=420)

        if "excel_bytes" in st.session_state and st.session_state["excel_bytes"]:
            st.success("Excel is ready.")
            st.download_button(
                label="Download Excel model",
                data=st.session_state["excel_bytes"],
                file_name=st.session_state.get("excel_name", "Restock_Model.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    if build_btn:
        st.session_state["last_error"] = ""
        st.session_state["excel_bytes"] = None
        st.session_state["stats"] = None
        st.session_state["full_preview"] = None
        st.session_state["excel_name"] = out_name.strip() or default_name

        if not (odoo_file and shop_file and pivot_file):
            st.session_state["last_error"] = "Please upload all 3 input files (Odoo Product Variant, Shopify CSV, Odoo Pivot)."
            st.rerun()

        try:
            full, days_window, stats = build_full_df(
                odoo_bytes=odoo_file.getvalue(),
                shopify_bytes=shop_file.getvalue(),
                shopify_filename=shop_file.name,
                pivot_bytes=pivot_file.getvalue(),
            )
            st.session_state["stats"] = stats
            st.session_state["full_preview"] = full.head(40)
            excel_bytes = build_excel_bytes(full, days_window, settings_overrides)
            st.session_state["excel_bytes"] = excel_bytes
            st.rerun()
        except Exception as e:
            st.session_state["last_error"] = f"Build failed: {e}"
            st.rerun()


if __name__ == "__main__":
    main()
